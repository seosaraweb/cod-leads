#!/usr/bin/env python3
import sys
import json
import io
import openpyxl
from openpyxl.styles import Font

orders = json.loads(sys.argv[1])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Colis"

headers = ['CODE SUIVI','DESTINATAIRE','TELEPHONE','ADRESSE','PRIX','VILLE','COMMENTAIRE','QUARTIER','PRODUIT','VALEUR DECLAREE']
widths  = [26, 23, 18, 18, 11, 20, 36, 9, 30, 19]

ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, name='Arial', size=10)

for col_idx, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

for o in orders:
    ws.append([
        o.get('order_ref',''),
        o.get('client_name',''),
        str(o.get('phone','')),
        o.get('address',''),
        float(o.get('price',0)) * int(o.get('quantity',1)),
        o.get('city',''),
        o.get('notes','') or '',
        '',
        o.get('product_name','') + ((' - ' + o['variant_label']) if o.get('variant_label') else '') + ((' x' + str(o['quantity'])) if int(o.get('quantity',1)) > 1 else ''),
        float(o.get('price',0)) * int(o.get('quantity',1)),
    ])

buf = io.BytesIO()
wb.save(buf)
sys.stdout.buffer.write(buf.getvalue())
